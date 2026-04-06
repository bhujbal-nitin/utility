import httpx
import asyncio

async def test_upload():
    # Attempt to upload a dummy file to the local proposal service
    url = "http://localhost:8002/api/proposal/upload"
    # We need a token. The user's token is in the metadata.
    token = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE3NzU3MTc1MTIsInN1YiI6IjRlYTQ5Y2IyLWJmNDctNDI4YS1iMDk4LWJkMDRjMmZhZWY2MiJ9.vj3hLt1_bxpJ5lFixo3sSbOrGZ8xPhiFUH75HETixVM"
    headers = {"Authorization": f"Bearer {token}"}
    
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Sr. No.")
    ws.cell(1, 2, "Process Name")
    ws.cell(2, 1, 1)
    ws.cell(2, 2, "Test Process")
    wb.save("test_discovery.xlsx")
    
    async with httpx.AsyncClient() as client:
        try:
            with open("test_discovery.xlsx", "rb") as f:
                files = {"file": ("test_discovery.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                resp = await client.post(url, headers=headers, files=files, timeout=120.0)
                print(f"Status: {resp.status_code}")
                print(f"Body: {resp.text}")
        except Exception as e:
            import traceback
            print(f"Request failed: {e}")
            traceback.print_exc()

if __name__ == "__main__":
    asyncio.run(test_upload())
