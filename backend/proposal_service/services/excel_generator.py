"""
excel_generator.py
------------------
Generates the 6-sheet Scope Commercials Excel.
Consolidated from scope_portal v8 logic.
"""

import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "BDD7EE"
C_WHITE      = "FFFFFF"
C_ORANGE     = "FFC000"
C_GREEN      = "70AD47"
C_LIGHT_GRAY = "F2F2F2"
C_ALT_ROW    = "EBF3FB"
C_YELLOW     = "FFFF99"

_COMPLEXITY_COLORS = {
    "Simple":  "E2EFDA",
    "Medium":  "FFF2CC",
    "Complex": "FCE4D6",
}


def _hdr_font(bold=True, size=10, color=C_WHITE):
    return Font(name="Arial", bold=bold, size=size, color=color)

def _body_font(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Sheet 1: Discovery ────────────────────────────────────────────────────────
def _build_discovery_sheet(wb_out, discovery_filepath):
    ws_out = wb_out.create_sheet("Discovery")
    wb_src = load_workbook(discovery_filepath, data_only=True)
    ws_src = wb_src.active

    all_rows = list(ws_src.iter_rows(values_only=True))
    if not all_rows:
        ws_out.freeze_panes = "A2"
        return

    header_row = list(all_rows[0])

    # Find and remove Efforts column index
    efforts_idx = None
    for i, h in enumerate(header_row):
        if h is not None:
            norm = str(h).lower().strip()
            if any(kw in norm for kw in ["professional service", "ps effort", "effort"]):
                efforts_idx = i
                break

    # Build column mapping: original col -> output col (skip efforts_idx)
    col_map_out = {}
    out_col = 0
    for i in range(len(header_row)):
        if i == efforts_idx:
            continue
        col_map_out[i] = out_col
        out_col += 1

    exec_time_out_col = out_col  # new column at end (0-based → +1 for Excel)

    # Write header row
    for orig_i, out_i in col_map_out.items():
        val = header_row[orig_i]
        if val is not None:
            c = ws_out.cell(row=1, column=out_i + 1, value=val)
            c.font = _hdr_font(); c.fill = _fill(C_DARK_BLUE)
            c.alignment = _center(); c.border = _border()

    exec_hdr = ws_out.cell(row=1, column=exec_time_out_col + 1,
                            value="Estimated Execution Time\nfor Bot (Min)\n(User to Fill)")
    exec_hdr.font = _hdr_font(); exec_hdr.fill = _fill(C_MID_BLUE)
    exec_hdr.alignment = _center(); exec_hdr.border = _border()

    # Write data rows
    for row_i, row in enumerate(all_rows[1:], start=2):
        row = list(row)
        while len(row) < len(header_row):
            row.append(None)

        for orig_i, out_i in col_map_out.items():
            val = row[orig_i] if orig_i < len(row) else None
            c = ws_out.cell(row=row_i, column=out_i + 1, value=val)
            c.font = _body_font(); c.alignment = _left(); c.border = _border()

        exec_c = ws_out.cell(row=row_i, column=exec_time_out_col + 1, value=None)
        exec_c.font = _body_font(); exec_c.fill = _fill(C_YELLOW)
        exec_c.alignment = _center(); exec_c.border = _border()

    for col_letter, dim in ws_src.column_dimensions.items():
        ws_out.column_dimensions[col_letter].width = dim.width or 18
    ws_out.column_dimensions[get_column_letter(exec_time_out_col + 1)].width = 22

    ws_out.freeze_panes = "A2"
    ws_out.row_dimensions[1].height = 45


# ── Sheet 2: Proposed Use Case ────────────────────────────────────────────────
_PUC_HEADERS = [
    "No", "Proposed Category", "Proposed Use case name",
    "Use Case Description\n(Daily Volume)", "Scope of Workflow/Steps of Process",
    "Other Info.", "Complexity",
    "Professional Services Efforts in Days (Delivery)",
    "Estimated Execution\nTime for Bot (Min)\n(User to Fill)",
    "Solution Mapping", "AI Agents Plugins",
]
_OTHER_INFO = (
    "Daily Volumes\nApplications/Systems\nNo. of screens\n"
    "No. of Fields\nException Scenario\nWorking Hours\nManual Staff"
)
_COL_WIDTHS_PUC = [6, 18, 35, 20, 22, 22, 14, 18, 18, 45, 12]


def _build_proposed_use_case(wb_out, use_cases):
    ws = wb_out.create_sheet("Proposed Use Case")
    ws.append([None] * 11)
    ws.row_dimensions[1].height = 8

    ws.append(_PUC_HEADERS)
    for col_idx in range(1, 12):
        c = ws.cell(row=2, column=col_idx)
        c.font = _hdr_font(size=9); c.fill = _fill(C_DARK_BLUE)
        c.alignment = _center(); c.border = _border()

    for i, uc in enumerate(use_cases):
        dr       = 3 + i
        row_fill = C_ALT_ROW if i % 2 == 0 else C_WHITE
        cx       = uc.get("complexity", "Medium")
        exec_t   = uc.get("estimated_exec_time", "") or ""

        ws.append([
            uc["sr_no"], "",
            uc["process_name"],
            f"{uc['daily_volume']} Cases/Daily",
            "Refer Discovery",
            _OTHER_INFO if i == 0 else "",
            cx,
            uc["ps_efforts"],
            exec_t,
            uc["solution_mapping"],
            uc["ai_plugins"],
        ])

        for col_idx in range(1, 12):
            c = ws.cell(row=dr, column=col_idx)
            c.font = _body_font(); c.border = _border()
            if col_idx == 7:
                c.fill = _fill(_COMPLEXITY_COLORS.get(cx, C_WHITE))
                c.font = _body_font(bold=True); c.alignment = _center()
            elif col_idx == 9:
                c.fill = _fill(C_YELLOW); c.alignment = _center()
            elif col_idx in (1, 8, 11):
                c.fill = _fill(row_fill); c.alignment = _center()
            else:
                c.fill = _fill(row_fill); c.alignment = _left()

    note_row  = 3 + len(use_cases) + 1
    last_data = 3 + len(use_cases) - 1
    ws.cell(row=note_row, column=1, value="Note").font = _body_font(bold=True)
    ws.cell(row=note_row, column=8,  value=f"=SUM(H3:H{last_data})").font = _body_font(bold=True)
    ws.cell(row=note_row, column=8).alignment = _center()
    ws.cell(row=note_row, column=11, value=f"=SUM(K3:K{last_data})").font = _body_font(bold=True)
    ws.cell(row=note_row, column=11).alignment = _center()

    notes = [
        "This is Proposed use cases tentatively for budgeting purposes",
        "Process Complexity will be finalised only after detailed requirement Gathering",
        "Efforts estimates may vary on complexity of process and will be communicated with necessary approvals",
        "Scope of work for each use case is tentative and will be finalised after detailed Requirement Gathering",
        "PS Efforts are as per Ref. Complexity Grid; any deviation in complexity parameters will have corresponding changes",
        "Estimated Execution Time for Bot (Min) is user-filled per use case; used to compute avg cycle time in Software sheet",
    ]
    for j, note in enumerate(notes, start=1):
        r = note_row + j
        ws.cell(row=r, column=1, value=j)
        ws.cell(row=r, column=2, value=note).font = _body_font(size=9)

    for col_idx, w in enumerate(_COL_WIDTHS_PUC, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 45


# ── Sheet 3: Use Case Volume ──────────────────────────────────────────────────
_VOL_HEADERS = [
    "Sr. No.", "Proposed Use Cases", "Volume (daily)",
    "Average Monthly Tickets for RPA", "Average Tickets Annually for RPA",
    "Documents Annually",
]
_COL_WIDTHS_VOL = [8, 45, 16, 24, 24, 20]


def _build_use_case_volume(wb_out, use_cases):
    ws = wb_out.create_sheet("Use Case Volume")
    ws.append(_VOL_HEADERS)
    for col_idx in range(1, 7):
        c = ws.cell(row=1, column=col_idx)
        c.font = _hdr_font(); c.fill = _fill(C_DARK_BLUE)
        c.alignment = _center(); c.border = _border()

    for i, uc in enumerate(use_cases):
        dr       = 2 + i
        row_fill = C_ALT_ROW if i % 2 == 0 else C_WHITE
        ws.append([
            uc["sr_no"], uc["process_name"], uc["daily_volume"],
            f"=C{dr}*30", f"=D{dr}*12", uc["docs_annually"],
        ])
        for col_idx in range(1, 7):
            c = ws.cell(row=dr, column=col_idx)
            c.font = _body_font(); c.fill = _fill(row_fill)
            c.border = _border()
            c.alignment = _left() if col_idx == 2 else _center()

    total_row = 2 + len(use_cases)
    last_data = total_row - 1
    ws.append(["Total", "", f"=SUM(C2:C{last_data})", f"=SUM(D2:D{last_data})",
               f"=SUM(E2:E{last_data})", f"=SUM(F2:F{last_data})"])
    for col_idx in range(1, 7):
        c = ws.cell(row=total_row, column=col_idx)
        c.font = _body_font(bold=True); c.fill = _fill(C_ORANGE)
        c.border = _border(); c.alignment = _center()

    for col_idx, w in enumerate(_COL_WIDTHS_VOL, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"


# ── Sheet 4: Ref Complexity Grid ──────────────────────────────────────────────
_REF_ROWS = [
    (1, "Definitions", "-",
     "Simple input (Excel/web), maximum simple binary decisions (Yes/No), no scanned/handwritten docs",
     "3-4 apps, moderate fields, attended automation may be needed, some exceptions",
     "5+ apps, multi-system handoffs, HITL, heavy IDP/AI/ML, mission-critical SLA"),
    (2, "# of Applications",      0.20, "1-2",       "3-4",     "5+"),
    (3, "# of fields",            0.25, "<50",        "51-75",   "76+"),
    (4, "# of Screens",           0.15, "<10",        "11-20",   "21+"),
    (5, "# of decision points",   0.10, "<3",         "4-7",     "8+"),
    (6, "Strict SLA",             0.10, "No",         "Yes",     "Critical"),
    (7, "Extent of Exceptions",   0.10, "<5",         "6-10",    "11+"),
    (8, "# of Validations/Rules", 0.10, "<10",        "11-20",   "21+"),
    (None, "Approx. Dev Efforts", None, "20-40 days", "45-65 days", "70-120 days"),
]
_REF_NOTES = [
    "Above Estimates are guidelines for Budgetary purpose only",
    "Actual Efforts may vary based on actual requirement understanding of each project / process",
    "Client will need to provide AS IS Process walkthrough during requirement gathering",
    "PS Efforts may vary depending on Solution Design & HyperAutomation components (RPA, OCR, ETL, API, AI, ML)",
]
_COMPLEXITY_HDR_COLORS = {
    "Simple Complexity":  "70AD47",
    "Medium Complexity":  "FFC000",
    "Complex Complexity": "FF0000",
}


def _build_ref_complexity_grid(wb_out):
    ws = wb_out.create_sheet("Ref Complexity Grid")
    ws.append([None] * 7)
    ws.row_dimensions[1].height = 8

    headers = ["#", "Header", "Weightage", "Simple Complexity", "Medium Complexity", "Complex Complexity"]
    ws.append(headers)
    for col_idx, hdr in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx)
        color = _COMPLEXITY_HDR_COLORS.get(hdr, C_DARK_BLUE)
        c.font = _hdr_font(); c.fill = _fill(color)
        c.alignment = _center(); c.border = _border()

    for i, row_data in enumerate(_REF_ROWS, start=3):
        ws.append(list(row_data))
        for col_idx in range(1, 7):
            c = ws.cell(row=i, column=col_idx)
            c.font = _body_font(); c.border = _border()
            c.alignment = _left() if col_idx == 2 else _center()
            c.fill = _fill(C_ALT_ROW if i % 2 == 0 else C_WHITE)

    note_start = 3 + len(_REF_ROWS) + 1
    ws.cell(row=note_start, column=1, value="Note").font = _body_font(bold=True)
    for j, note in enumerate(_REF_NOTES, start=1):
        r = note_start + j
        ws.cell(row=r, column=1, value=j)
        ws.cell(row=r, column=2, value=note).font = _body_font(size=9)

    col_widths = [6, 30, 12, 32, 32, 32]
    for col_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 30


# ── Production RAM calculation ─────────────────────────────────────────────────
def _round_to_standard_ram(gb: float) -> int:
    """Round up to the nearest standard RAM size: 8, 16, 32, 64, 128, 256, 512 GB."""
    standards = [8, 16, 32, 64, 128, 256, 512]
    for s in standards:
        if gb <= s:
            return s
    return math.ceil(gb / 128) * 128


def _calc_production_ram(num_bots: int = 18, num_cores: int = 6) -> int:
    """
    Production environment RAM calculation:
    - 1 bot = 4 GB RAM
    - 5 GB for OS
    - 30% buffer
    - Core-to-RAM ratio = 1:6
    Round up to standard RAM size.
    """
    bot_ram    = num_bots * 4
    os_ram     = 5
    base_ram   = bot_ram + os_ram
    with_buffer = base_ram * 1.3

    core_ratio_ram = num_cores * 6        # 1:6 ratio
    needed = max(with_buffer, core_ratio_ram)
    return _round_to_standard_ram(needed)


# ── Sheet 5: Software ─────────────────────────────────────────────────────────
def _build_software_sheet(wb_out, use_cases, work_days=1, total_bots=18, sys_hours=16.0):
    ws = wb_out.create_sheet("Software")
    n_uc          = len(use_cases)
    vol_total_row = n_uc + 2
    puc_note_row  = n_uc + 4

    # ── Col G: Exec time reference table ────────
    g_hdr = ws.cell(row=1, column=7, value="Exec Time per Use Case (Min)\n[Update from Discovery]")
    g_hdr.font = _hdr_font(size=9); g_hdr.fill = _fill(C_MID_BLUE)
    g_hdr.alignment = _center(); g_hdr.border = _border()
    ws.row_dimensions[1].height = 30

    for i, uc in enumerate(use_cases):
        r   = 2 + i
        val = uc.get("estimated_exec_time", 0) or 0
        c   = ws.cell(row=r, column=7, value=val)
        c.font = _body_font(size=9); c.fill = _fill(C_YELLOW)
        c.alignment = _center(); c.border = _border()

    exec_r_start = 2
    exec_r_end   = 1 + n_uc
    avg_formula  = f"=IFERROR(AVERAGEIF(G{exec_r_start}:G{exec_r_end},\">0\"),5)"

    def hrow(row_num, vals, fill=C_MID_BLUE):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=ci, value=v)
            c.font = _hdr_font(); c.fill = _fill(fill)
            c.alignment = _center(); c.border = _border()

    def drow(row_num, vals, fill=C_WHITE, bold=False, num_fmt=None):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=ci, value=v)
            c.font = _body_font(bold=bold); c.fill = _fill(fill)
            c.border = _border()
            c.alignment = _left() if ci == 2 else _center()
            if num_fmt and ci == 3:
                c.number_format = num_fmt

    hrow(1, [1, "Description", "Value", "Remark/Assumption", None, None])
    drow(2, [1.1, "System Available Hours per Day", sys_hours, f"Enter hours the system is available per day (default: {sys_hours})"])
    ws.cell(row=2, column=3).fill = _fill(C_YELLOW)

    drow(3, [1.2, "No. of Actual Working Days per Month", work_days, "Worst case – Holidays & System Availability"])
    drow(4, [1.3, "Lowest Available Daily System Time (Min)", "=C2*60", "=C2*60"])
    drow(5, [1.4, "BOTs Total Time Available (Min) — for 1 BOT", "=C4*C3*1", "Assumed one bot for calculation"])

    hrow(7, [2, "Description", "Value", "Remark/Assumption", None, None])
    drow(8,  [2.1, "Estimated Cycle Time Per Ticket (Min)", avg_formula,
              "AVERAGE of all use case estimated bot execution times (update Col G above)"])
    drow(9,  [2.2, "Ticket/Request Per Day (Volume)",
              f"='Use Case Volume'!C{vol_total_row}", None])
    drow(10, [2.3, "Time Required for Daily Processing (Min)", "=C8*C9", None])

    # Note: Using standard logic C10/C5 for bot count calculation
    drow(11, [1.5, "Total BOTs Required", "=ROUNDUP(C10/C5, 0)", "Total Bots Required"])
    ws.cell(row=11, column=3).font = _body_font(bold=True)
    ws.cell(row=11, column=3).fill = _fill(C_LIGHT_BLUE)

    metrics = [
        ("No. of BOTs",                           "=C11",      None),
        ("Productive Capacity",                    "=C5/C10",  "0.00%"),
        ("Expected Total Production Capacity/Day", "=C13*C9",  None),
        ("Expected Cases/BOT/Day",                 "=C9/C11",   None),
        ("BOT Utilisation",                        "=C10/C5",  "0.00%"),
    ]
    for r_off, (label, formula, num_fmt) in enumerate(metrics, start=13):
        row_fill = C_ALT_ROW if r_off % 2 == 0 else C_WHITE
        for ci, v in enumerate([None, label, formula, None], 1):
            c = ws.cell(row=r_off, column=ci, value=v)
            c.font = _body_font(); c.fill = _fill(row_fill); c.border = _border()
            c.alignment = _left() if ci == 2 else _center()
            if num_fmt and ci == 3:
                c.number_format = num_fmt

    hrow(20, [None, "Commercials", None, None, None, None])
    hrow(21, ["No.", "Annual Subscription Based License Line Item",
              "No. of Units", "Cost / Unit / Year (INR)", "Total Cost (INR)", None],
         fill=C_DARK_BLUE)
    drow(22, [1, "AutomationEdge RPA Advanced Unassisted Bot", "=C13", 250000, "=C22*D22"])
    drow(23, [2, "DocEdge", f"='Use Case Volume'!F{vol_total_row}", 2, "=C23*D23"])
    drow(24, [3, "Agentic AI Plugins (LLM connector + Classifier + Summariser + RAG + AI Master Conductor)",
              f"='Proposed Use Case'!K{puc_note_row}", 300000, "=C24*D24"])
    drow(25, ["Total", None, None, None, "=SUM(E22:E24)"], fill=C_ORANGE, bold=True)

    sw_notes = [
        "Above Estimates are guidelines for Budgetary purpose only",
        "Actual Efforts may vary based on actual requirement understanding of each project / process",
        "Client will need to provide AS IS Process walkthrough during requirement gathering",
        "PS Efforts may vary depending on Solution Design & HyperAutomation components (RPA, OCR, ETL, API, AI, ML)",
        "Above estimates are as per Ref. Complexity Grid; any deviation may have corresponding change in PS Efforts",
        "Estimated Cycle Time Per Ticket = AVERAGE of Exec Time column G (update values from Discovery sheet yellow cells)",
        "System Available Hours (C2, yellow) is user-editable; change it to reflect actual system availability.",
    ]
    ws.cell(row=27, column=1, value="Note").font = _body_font(bold=True)
    for j, note in enumerate(sw_notes, start=1):
        ws.cell(row=27+j, column=1, value=j)
        ws.cell(row=27+j, column=2, value=note).font = _body_font(size=9)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["G"].width = 24


# ── Sheet 6: Hardware ─────────────────────────────────────────────────────────
def _build_hardware_sheet(wb_out, use_cases: list[dict] = None, total_bots=18, num_cores=6):
    ws = wb_out.create_sheet("Hardware")

    # Determine total bots from use_cases context
    prod_bots  = total_bots
    prod_cores = num_cores
    prod_ram   = _calc_production_ram(prod_bots, prod_cores)

    sections = [
        ("Production Environment (Onpremise)",
         ["No. of Servers", "Applications / Module", "Server", "CPU", "Core", "RAM (GB)", "HD (GB)", "Operating System", "DB", "Web Server"],
         [[2, "AutomationEdge Processing Server incl. Active MQ & PostgreSQL Database",
           "VM", 3, prod_cores, prod_ram, 500, "MS Windows Server 2022/2023 – 64 bit", "PostgreSQL (Default)", "Apache Tomcat"]]),
        ("UAT Environment (Onpremise)",
         ["No. of Servers", "Applications", "Server", "CPU", "Core", "RAM (GB)", "HD (GB)", "Operating System", "DB", "Web Server"],
         [[1, "AE Main Server, DocEdge Server, Active MQ, PostgreSQL Database & Processing Server",
           "VM", 1, 4, 16, 500, "MS Windows Server 2022/2023 – 64 bit", "PostgreSQL (Default)", "Apache Tomcat"]]),
        ("Development Environment (Offshore Desktop Systems)",
         ["No. of Desktops", "Applications", "Type", "CPU", "Core", "RAM (GB)", "HD (GB)", "Operating System", "DB", "Web Server"],
         [[4, "Desktop Development Machine for Chatbot / Script Development",
           "Desktop / VM with Remote Access", 1, 4, 8, 500, "Windows 7 Professional – 64 bit", "PostgreSQL (Default)", "NA"]]),
    ]

    cur = 1
    for title, hdrs, rows in sections:
        ws.cell(row=cur, column=1, value=title)
        ws.cell(row=cur, column=1).font = _hdr_font(size=11)
        ws.cell(row=cur, column=1).fill = _fill(C_DARK_BLUE)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=10)
        cur += 1
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=cur, column=ci, value=h)
            c.font = _hdr_font(); c.fill = _fill(C_MID_BLUE)
            c.alignment = _center(); c.border = _border()
        cur += 1
        for dr in rows:
            for ci, v in enumerate(dr, 1):
                c = ws.cell(row=cur, column=ci, value=v)
                c.font = _body_font(); c.fill = _fill(C_LIGHT_GRAY)
                c.border = _border()
                c.alignment = _left() if ci in (2, 8, 9, 10) else _center()
            cur += 1
        cur += 1

    # RAM formula note for Production
    ws.cell(row=cur, column=1,
            value=f"Production RAM formula: ({prod_bots} bots × 4 GB) + 5 GB OS = {prod_bots*4+5} GB base "
                  f"→ + 30% buffer = {(prod_bots*4+5)*1.3:.1f} GB "
                  f"→ Core ratio ({prod_cores}×6={prod_cores*6} GB) "
                  f"→ Rounded to standard: {prod_ram} GB")
    ws.cell(row=cur, column=1).font = _body_font(size=8, color="595959")
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=10)
    cur += 2

    notes = [
        "Server/Desktop count may increase/decrease based on actual requirements and volumes.",
        "All PC/VM/Desktops should be in the same domain.",
        "AutomationEdge includes PostgreSQL; Oracle or MSSQL can be used at additional cost.",
        "Connectivity of target systems from the AutomationEdge server must be provisioned for GUI and REST APIs.",
        "Power user rights required on target systems and provided VM/PC/Desktops.",
        "Hardware procurement, installation, and maintenance are the client's responsibility.",
        "Development is performed on the UAT instance; deployment to production follows client UAT approval.",
        f"Production RAM sized for {prod_bots} BOTs × 4 GB + 5 GB OS + 36% spare, 1:6 core-to-RAM ratio, rounded to {prod_ram} GB standard.",
    ]
    ws.cell(row=cur, column=1, value="Notes").font = _body_font(bold=True)
    cur += 1
    for j, note in enumerate(notes, 1):
        ws.cell(row=cur, column=1, value=f"{j}. {note}").font = _body_font(size=9)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=10)
        cur += 1

    # Column Widths
    col_widths_hw = [10, 55, 18, 8, 8, 10, 10, 38, 20, 18]
    for ci, w in enumerate(col_widths_hw, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def generate_scope_excel(output_path, discovery_filepath, use_cases, work_days=1, total_bots=18, cycle_time=5, num_cpu=1, num_cores=4, sys_hours=16.0, hw_data=None):
    wb = Workbook()
    wb.remove(wb.active)
    _build_discovery_sheet(wb, discovery_filepath)
    _build_proposed_use_case(wb, use_cases)
    _build_use_case_volume(wb, use_cases)
    _build_ref_complexity_grid(wb)
    _build_software_sheet(wb, use_cases, work_days=work_days, total_bots=total_bots, sys_hours=sys_hours)
    _build_hardware_sheet(wb, use_cases, total_bots=total_bots, num_cores=num_cores)
    wb.save(output_path)
    return output_path
