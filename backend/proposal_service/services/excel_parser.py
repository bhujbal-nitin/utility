"""
excel_parser.py
---------------
Header-driven, column-order-agnostic parser for Discovery sheets.
Changes v7:
  - Removed ps_efforts (Efforts column removed from discovery template)
  - Added estimated_exec_time (user fills: estimated bot execution time per ticket in minutes)
"""

import re
from openpyxl import load_workbook
from proposal_service.services.volume_converter import to_daily


_FIELD_PATTERNS = [
    ("sr_no",               ["sr.no", "sr no", "srno", "serial"]),
    ("process_name",        ["process name", "process"]),
    ("sme",                 ["sme", "user name", "username"]),
    ("apps",                ["application", "no of app", "apps"]),
    ("nature",              ["nature of process", "nature"]),
    ("user_profile",        ["user profile", "profile"]),
    ("num_users",           ["no.of user", "no of user", "num user", "number of user"]),
    ("frequency",           ["frequency"]),
    ("process_time",        ["process completion time", "step wise time", "completion time", "process time"]),
    ("tat",                 ["tat", "sla"]),
    ("work_hours",          ["working hour", "work hour"]),
    ("trigger",             ["trigger", "how input", "how activity"]),
    ("idp",                 ["idp"]),
    ("docs_annually",       ["document annually", "docs annually", "documents annually"]),
    # ps_efforts intentionally removed from discovery (Requirement #1)
    ("estimated_exec_time", ["estimated execution time", "execution time for bot", "bot execution time",
                              "estimated exec", "exec time per", "execution time"]),
    ("raw_volume",          ["volume"]),
    ("input_files",         ["no.of input", "no of input", "input file"]),
    ("input_type",          ["type of input"]),
    ("ocr_nlp",             ["ocr", "nlp", "machine learning"]),
    ("input_desc",          ["description of data in input", "input description", "description of input"]),
    ("output_files",        ["no of output", "no.of output", "output file"]),
    ("output_type",         ["type of output"]),
    ("output_desc",         ["description of data in output", "output description", "description of output"]),
    ("pain_areas",          ["pain area"]),
    ("summary",             ["process summary", "summary"]),
    ("remarks",             ["remark"]),
]


def _normalise(text):
    t = str(text or "").lower()
    t = re.sub(r"[^a-z0-9 .]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def _build_col_map(header_row):
    col_map = {}
    normalised_headers = [_normalise(h) for h in header_row]
    for field_key, patterns in _FIELD_PATTERNS:
        for col_idx, nh in enumerate(normalised_headers):
            if any(pat in nh for pat in patterns) and field_key not in col_map:
                col_map[field_key] = col_idx
                break
    return col_map


def _get(row, col_map, field, default=""):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    return str(val).strip() if val is not None else default


def _get_int(row, col_map, field, default=0):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    if val is None:
        return default
    try:
        return int(float(str(val).strip().replace(",", "")))
    except (ValueError, TypeError):
        return default


def _get_float(row, col_map, field, default=0.0):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    if val is None:
        return default
    try:
        return float(str(val).strip().replace(",", ""))
    except (ValueError, TypeError):
        return default


def _is_sr_no(val):
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        float(str(val).strip())
        return True
    except ValueError:
        return False


def _find_header_row(ws):
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx > 10:
            break
        first = _normalise(row[0] if row else "")
        if "sr" in first or "no" in first:
            return row_idx
    return 0


def parse_discovery_sheet(filepath):
    wb = load_workbook(filepath, data_only=True)

    ws = None
    for sheet_name in wb.sheetnames:
        candidate = wb[sheet_name]
        first_cell = str(candidate.cell(1, 1).value or "").lower()
        if any(k in first_cell for k in ["sr", "no", "process", "1"]):
            ws = candidate
            break
    if ws is None:
        ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    hdr_idx = _find_header_row(ws)
    header_row = all_rows[hdr_idx]
    col_map = _build_col_map(header_row)

    records = []
    for row in all_rows[hdr_idx + 1:]:
        if not row or all(v is None for v in row):
            continue

        sr_val = row[col_map.get("sr_no", 0)] if col_map.get("sr_no", 0) < len(row) else None
        if not _is_sr_no(sr_val):
            continue

        try:
            sr_no = int(float(str(sr_val).strip()))
        except (ValueError, TypeError):
            continue

        raw_volume = _get(row, col_map, "raw_volume")
        daily_vol  = to_daily(raw_volume)

        records.append({
            "sr_no":                sr_no,
            "process_name":         _get(row, col_map, "process_name"),
            "sme":                  _get(row, col_map, "sme"),
            "apps":                 _get(row, col_map, "apps"),
            "user_profile":         _get(row, col_map, "user_profile"),
            "num_users":            _get(row, col_map, "num_users"),
            "frequency":            _get(row, col_map, "frequency"),
            "process_time":         _get(row, col_map, "process_time"),
            "work_hours":           _get(row, col_map, "work_hours"),
            "idp":                  _get(row, col_map, "idp"),
            "docs_annually":        _get_int(row, col_map, "docs_annually"),
            # ps_efforts removed from discovery sheet
            "estimated_exec_time":  _get_float(row, col_map, "estimated_exec_time", 0.0),
            "raw_volume":           raw_volume,
            "daily_volume":         daily_vol,
            "tat":                  _get(row, col_map, "tat"),
            "nature":               _get(row, col_map, "nature"),
            "trigger":              _get(row, col_map, "trigger"),
            "input_files":          _get(row, col_map, "input_files"),
            "input_type":           _get(row, col_map, "input_type"),
            "ocr_nlp":              _get(row, col_map, "ocr_nlp"),
            "input_desc":           _get(row, col_map, "input_desc"),
            "output_files":         _get(row, col_map, "output_files"),
            "output_type":          _get(row, col_map, "output_type"),
            "output_desc":          _get(row, col_map, "output_desc"),
            "pain_areas":           _get(row, col_map, "pain_areas"),
            "summary":              _get(row, col_map, "summary"),
            "remarks":              _get(row, col_map, "remarks"),
            # AI-filled placeholders
            "complexity":           "",
            "ps_efforts":           0,
            "solution_mapping":     "",
            "ai_plugins":           1,
        })

    return records
